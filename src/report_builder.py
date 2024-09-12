import datetime
from collections import defaultdict
from contextlib import contextmanager

from openpyxl import Workbook
from openpyxl import load_workbook
import getpass


class ReportBuilder:

    @staticmethod
    @contextmanager
    def read_rexcel(excel_path: str):
        wb = load_workbook(filename=excel_path, read_only=True)
        ws = wb.worksheets[0]
        try:
            yield ws
        finally:
            wb.close()

    @staticmethod
    @contextmanager
    def write_excel():
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title='Luntry')
        try:
            yield ws
        finally:
            wb.save(f'Уязвимости-{datetime.datetime.today().strftime("%d.%m.%Y-%Hh.%Mm")}.xlsx')
            wb.close()

    @staticmethod
    def update_db(excel_path: str, reports_by_image: dict[str, list[dict[str, str]]], reports_with_cve: dict[str, set]):
        current_date = datetime.datetime.today().strftime('%d/%m/%Y')
        with ReportBuilder.read_rexcel(excel_path) as ws_read, ReportBuilder.write_excel() as ws_save:

            # обновляем текущий отчет
            cve_in_db = defaultdict(set)
            extra_append = []
            for i, row in enumerate(ws_read.rows):
                if i == 0:
                    columns = [cell.value for cell in row]
                    ws_save.append(columns)
                    continue

                values = [str(cell.value) for cell in row if cell.value is not None]

                # получаем значения
                image_name = values[14]
                cve_id = values[0]
                state_pos = 16
                image_id_pos = 11

                cve_in_db[image_name].add(cve_id)

                if image_name in reports_with_cve:
                    if cve_id not in reports_with_cve[image_name]:
                        values[state_pos] = 'Closed'
                        ws_save.append(values)
                    else:
                        need_reports = reports_by_image[image_name]
                        found_report = next(report for report in need_reports if report["ID"] == cve_id)
                        found_report = list(map(str, found_report.values()))

                        to_append = [current_date, image_name, getpass.getuser(), 'Active']
                        found_report.extend(to_append)
                        if values[image_id_pos] == found_report[image_id_pos]:
                            if len(values) == 18:
                                found_report.append(str(values[-1]))
                            ws_save.append(found_report)
                        else:
                            if values[state_pos] == 'Closed':
                                extra_append.append(found_report)
                            ws_save.append(values)

                else:
                    ws_save.append(values)

            # добавим новые уязвимости
            for image in reports_by_image:
                to_append = [current_date, image, getpass.getuser(), 'Active']

                # образа нет в текущей БД
                if image not in cve_in_db:
                    for report in reports_by_image[image]:
                        values = list(map(str, report.values()))
                        values.extend(to_append)
                        ws_save.append(values)
                else:
                    existing_cve = cve_in_db[image]
                    new_cve = reports_with_cve[image] - existing_cve

                    for report in reports_by_image[image]:
                        if report['ID'] in new_cve:
                            values = list(map(str, report.values()))
                            values.extend(to_append)
                            ws_save.append(values)

            # добавим "старые" новые уязвимости
            for row in extra_append:
                ws_save.append(row)
